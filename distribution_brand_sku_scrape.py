#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scrape all SKUs from every brand on All-Stars Distribution.

Flow:
  1. Login (same as update_prices_and_sheet.py / distribution_cost_scrape.py)
  2. Open /en/manufacturers and collect brand listing URLs
  3. For each brand, paginate through the product grid and read SKU from each card
  4. Save Brand + SKU (+ URLs) to Excel in this project folder

Examples:
  python distribution_brand_sku_scrape.py
  python distribution_brand_sku_scrape.py --limit-brands 3
  python distribution_brand_sku_scrape.py --brand-url "https://www.all-stars-distribution.com/en/76_alpha-competition"
  python distribution_brand_sku_scrape.py --output distribution_brand_skus.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from datetime import datetime
from typing import Callable, Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin, urlparse

from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from seleniumbase import Driver

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from distribution_cost_scrape import (
    CARD_SKU_XPATHS,
    DISTRIBUTION_EMAIL,
    DISTRIBUTION_LOGIN_URL,
    DISTRIBUTION_PASSWORD,
    LIST_WAIT_SEC,
    PAGINATION_CLICK_WAIT_SEC,
    PAGINATION_UL_XPATH,
    RESULT_LI_XPATH,
    RESULT_LIST_UL_XPATHS,
    _extract_card_sku,
    _wait_for_result_cards,
    login_distribution_site,
)

DISTRIBUTION_BASE = "https://www.all-stars-distribution.com"
MANUFACTURERS_URL = f"{DISTRIBUTION_BASE}/en/manufacturers"
BRAND_HREF_RE = re.compile(r"/en/\d+_[\w-]+/?$", re.I)

MANUFACTURER_LINK_XPATHS = (
    '//*[@id="center_column"]//h3/a',
    '//*[@id="center_column"]//a[contains(@href, "_")]',
    '//ul[contains(@class,"manufacturer")]//a',
    '//div[contains(@class,"manufacturer")]//a',
)

DEFAULT_OUTPUT = os.path.join(
    SCRIPT_DIR,
    f"distribution_brand_skus_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
)

EXCEL_HEADERS = (
    "brand_name",
    "sku",
    "brand_url",
    "listing_page_url",
    "page_number",
)


def _slug_to_brand_name(slug: str) -> str:
    """Fallback brand label from URL slug e.g. alpha-competition -> Alpha Competition."""
    text = slug.replace("-", " ").replace("_", " ")
    return " ".join(part.capitalize() for part in text.split())


def _brand_name_from_url(url: str) -> str:
    path = urlparse(url).path.rstrip("/")
    slug = path.split("/")[-1]
    if "_" in slug:
        slug = slug.split("_", 1)[1]
    return _slug_to_brand_name(slug)


def _normalize_brand_url(url: str) -> str:
    parsed = urlparse(url.split("?")[0].split("#")[0])
    return f"{DISTRIBUTION_BASE}{parsed.path.rstrip('/')}"


def _is_brand_listing_url(url: str) -> bool:
    path = urlparse(url).path
    return bool(BRAND_HREF_RE.search(path))


def _wait_for_listing_or_empty(driver, timeout: float = LIST_WAIT_SEC) -> str:
    """
    Return 'products' when product cards are visible, 'empty' when page has no listing,
    'timeout' otherwise.
    """
    end = time.time() + timeout
    while time.time() < end:
        try:
            cards = driver.find_elements(By.XPATH, RESULT_LI_XPATH)
            if cards:
                return "products"
        except Exception:
            pass
        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "no products" in body or "0 product" in body:
                return "empty"
        except Exception:
            pass
        for xp in RESULT_LIST_UL_XPATHS:
            try:
                ul = driver.find_element(By.XPATH, xp)
                if ul.is_displayed():
                    if ul.find_elements(By.TAG_NAME, "li"):
                        return "products"
            except Exception:
                continue
        time.sleep(0.4)
    return "timeout"


def scrape_skus_on_current_page(driver) -> List[str]:
    skus: List[str] = []
    seen: Set[str] = set()
    try:
        cards = _wait_for_result_cards(driver)
    except Exception:
        cards = driver.find_elements(By.XPATH, RESULT_LI_XPATH)

    for card in cards:
        sku = _extract_card_sku(card)
        if not sku:
            for xp in CARD_SKU_XPATHS:
                try:
                    sku = card.find_element(By.XPATH, xp).text.strip()
                    if sku:
                        break
                except Exception:
                    continue
        sku = str(sku or "").strip()
        if sku and sku not in seen:
            seen.add(sku)
            skus.append(sku)
    return skus


def _scroll_pagination_into_view(driver) -> None:
    try:
        pag = driver.find_element(By.XPATH, PAGINATION_UL_XPATH)
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
            pag,
        )
        time.sleep(0.35)
    except Exception:
        pass


def go_to_next_listing_page(driver, current_page: int) -> bool:
    """Click pagination link for page current_page + 1."""
    next_page = current_page + 1
    _scroll_pagination_into_view(driver)
    try:
        pagination = driver.find_element(By.XPATH, PAGINATION_UL_XPATH)
    except Exception:
        return False

    for li in pagination.find_elements(By.TAG_NAME, "li"):
        try:
            a = li.find_element(By.TAG_NAME, "a")
            txt = (a.text or "").strip()
            if txt.isdigit() and int(txt) == next_page:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
                    a,
                )
                time.sleep(0.2)
                try:
                    a.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", a)
                time.sleep(PAGINATION_CLICK_WAIT_SEC)
                try:
                    _wait_for_result_cards(driver, timeout=LIST_WAIT_SEC)
                except Exception:
                    pass
                return True
        except Exception:
            continue
    return False


def read_brand_title(driver) -> str:
    for xp in (
        '//*[@id="center_column"]//h1',
        '//*[@id="center_column"]//h2',
        "//h1",
    ):
        try:
            text = driver.find_element(By.XPATH, xp).text.strip()
            if text and text.lower() not in {"brands", "manufacturers", "login"}:
                return text
        except Exception:
            continue
    return ""


def scrape_brand(
    driver,
    brand_url: str,
    brand_name: str,
    log: Callable[[str], None],
) -> List[Dict[str, str]]:
    brand_url = _normalize_brand_url(brand_url)
    rows: List[Dict[str, str]] = []
    driver.get(brand_url)
    time.sleep(1.5)

    page_name = read_brand_title(driver)
    if page_name:
        brand_name = page_name

    page = 1
    while True:
        state = _wait_for_listing_or_empty(driver)
        listing_url = driver.current_url

        if state == "empty":
            log(f"  {brand_name}: no products")
            break
        if state == "timeout" and page == 1:
            log(f"  {brand_name}: listing not found / timeout")
            break
        if state == "timeout":
            log(f"  {brand_name}: page {page} timeout — stopping pagination")
            break

        skus = scrape_skus_on_current_page(driver)
        log(f"  {brand_name} page {page}: {len(skus)} SKU(s)")
        for sku in skus:
            rows.append(
                {
                    "brand_name": brand_name,
                    "sku": sku,
                    "brand_url": brand_url,
                    "listing_page_url": listing_url,
                    "page_number": str(page),
                }
            )

        if not go_to_next_listing_page(driver, page):
            break
        page += 1
        if page > 500:
            log(f"  {brand_name}: safety stop at page 500")
            break

    return rows


def collect_brand_links(driver) -> List[Tuple[str, str]]:
    driver.get(MANUFACTURERS_URL)
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="center_column"]'))
    )
    time.sleep(1.5)

    found: Dict[str, str] = {}
    for xp in MANUFACTURER_LINK_XPATHS:
        for a in driver.find_elements(By.XPATH, xp):
            href = (a.get_attribute("href") or "").strip()
            if not href:
                continue
            full = urljoin(DISTRIBUTION_BASE, href.split("?")[0].split("#")[0])
            if not _is_brand_listing_url(full):
                continue
            norm = _normalize_brand_url(full)
            name = (a.text or "").strip() or _brand_name_from_url(norm)
            if norm not in found:
                found[norm] = name

    brands = sorted(found.items(), key=lambda x: x[1].lower())
    return [(url, name) for url, name in brands]


def save_excel(path: str, rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "distribution_skus"
    ws.append(list(EXCEL_HEADERS))
    for row in rows:
        ws.append([row.get(h, "") for h in EXCEL_HEADERS])
    wb.save(path)


def append_checkpoint(path: str, rows: List[Dict[str, str]]) -> None:
    """Append rows to checkpoint Excel (creates file with header if missing)."""
    if not rows:
        return
    if os.path.isfile(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "distribution_skus"
        ws.append(list(EXCEL_HEADERS))
    for row in rows:
        ws.append([row.get(h, "") for h in EXCEL_HEADERS])
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Scrape all brand SKUs from All-Stars Distribution into Excel.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Excel output path (default: distribution_brand_skus_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--limit-brands",
        type=int,
        default=0,
        help="Process only first N brands (0 = all)",
    )
    parser.add_argument(
        "--brand-url",
        action="append",
        default=[],
        help="Scrape only this brand listing URL (repeatable)",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Chrome headless",
    )
    args = parser.parse_args()

    output_path = os.path.abspath(args.output)
    os.makedirs(os.path.dirname(output_path) or SCRIPT_DIR, exist_ok=True)

    print("=" * 70)
    print("All-Stars Distribution — brand SKU scrape")
    print("=" * 70)
    print(f"Manufacturers: {MANUFACTURERS_URL}")
    print(f"Output:        {output_path}")
    print()

    driver = Driver(uc=True, headless=args.headless, incognito=True)
    driver.maximize_window()
    all_rows: List[Dict[str, str]] = []

    try:
        print("Logging in…")
        if not login_distribution_site(driver, DISTRIBUTION_EMAIL, DISTRIBUTION_PASSWORD):
            print("[ERROR] Distribution login failed.")
            return 1
        print("[OK] Logged in.\n")

        if args.brand_url:
            brands = [
                (_normalize_brand_url(u), _brand_name_from_url(u))
                for u in args.brand_url
            ]
        else:
            print("Collecting brand links from manufacturers page…")
            brands = collect_brand_links(driver)
            print(f"[OK] Found {len(brands)} brand(s).\n")

        if args.limit_brands > 0:
            brands = brands[: args.limit_brands]
            print(f"Limited to {len(brands)} brand(s).\n")

        total_skus = 0
        for idx, (brand_url, brand_name) in enumerate(brands, start=1):
            print(f"[{idx}/{len(brands)}] {brand_name}")
            print(f"  URL: {brand_url}")
            rows = scrape_brand(
                driver,
                brand_url,
                brand_name,
                log=lambda msg: print(msg),
            )
            all_rows.extend(rows)
            total_skus += len(rows)
            append_checkpoint(output_path, rows)
            print(f"  Total so far: {total_skus} SKU row(s)\n")

        save_excel(output_path, all_rows)
        unique_skus = len({(r["brand_name"], r["sku"]) for r in all_rows})
        print("=" * 70)
        print(f"Done. Brands: {len(brands)}")
        print(f"SKU rows: {len(all_rows)} ({unique_skus} unique brand+SKU pairs)")
        print(f"Saved: {output_path}")
        print("=" * 70)
        return 0
    finally:
        driver.quit()


if __name__ == "__main__":
    raise SystemExit(main())
