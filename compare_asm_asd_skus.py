#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare SKUs: ASM (Google Sheet) vs ASD (Distribution Excel scrape).

Creates two tabs on the ASM spreadsheet:
  • asm_not_in_asd  — products on ASM (motorsport) missing from ASD (distribution)
  • asd_not_in_asm  — products on ASD missing from ASM

ASM sheet: set GOOGLE_SPREADSHEET_ID in your .env file.

ASD source: Excel from distribution_brand_sku_scrape.py (columns: brand_name, sku, …)

Examples:
  python compare_asm_asd_skus.py
  python compare_asm_asd_skus.py --asd-xlsx distribution_brand_skus.xlsx
  python compare_asm_asd_skus.py --write-csv
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import random
import sys
import time
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError
from gspread.exceptions import APIError
from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASM_SPREADSHEET_ID = os.environ.get("GOOGLE_SPREADSHEET_ID", "")
ASM_PRODUCTS_TAB = "products"
TAB_ASM_NOT_IN_ASD = "asm_not_in_asd"
TAB_ASD_NOT_IN_ASM = "asd_not_in_asm"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
CREDENTIALS_FILE = os.path.join(SCRIPT_DIR, "credentials.json")

ASM_NOT_HEADERS = [
    "sku",
    "Reference",
    "Brand1",
    "Title",
    "product_id",
    "Price",
    "Availability",
    "url",
]

ASD_NOT_HEADERS = [
    "sku",
    "brand_name",
    "brand_url",
    "listing_page_url",
    "page_number",
]


def safe_gs_call(func, *args, max_retries=6, base_delay=1.25, **kwargs):
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except (HttpError, APIError) as e:
            status = None
            try:
                status = getattr(getattr(e, "resp", None), "status", None)
            except Exception:
                status = None
            msg = str(e).lower()
            retryable = status in (429, 500, 503) or ("quota" in msg) or ("rate" in msg and "limit" in msg)
            if not retryable or attempt >= max_retries - 1:
                raise
            time.sleep((base_delay * (2**attempt)) + random.uniform(0.0, 0.6))


def normalize_sku(raw: str) -> str:
    """Lowercase match key: strip #/spaces, then lowercase (R06-10-2 == r06-10-2)."""
    return str(raw or "").strip().lstrip("#").strip().lower()


def find_latest_asd_xlsx(explicit: str = "") -> str:
    if explicit and os.path.isfile(explicit):
        return os.path.abspath(explicit)
    patterns = [
        os.path.join(SCRIPT_DIR, "distribution_brand_skus*.xlsx"),
        os.path.join(SCRIPT_DIR, "distribution_brand_skus_test.xlsx"),
    ]
    candidates: List[str] = []
    for pattern in patterns:
        candidates.extend(glob.glob(pattern))
    if not candidates:
        raise FileNotFoundError(
            "No ASD Excel found. Run:\n"
            "  python distribution_brand_sku_scrape.py "
            f'--output "{os.path.join(SCRIPT_DIR, "distribution_brand_skus.xlsx")}"'
        )
    return max(candidates, key=os.path.getmtime)


def load_asd_rows(xlsx_path: str) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().lower() for c in next(rows_iter, [])]
    if not header:
        wb.close()
        return []

    def col_idx(*names: str) -> int:
        for name in names:
            if name in header:
                return header.index(name)
        return -1

    sku_i = col_idx("sku")
    brand_i = col_idx("brand_name", "brand")
    url_i = col_idx("brand_url")
    page_url_i = col_idx("listing_page_url")
    page_num_i = col_idx("page_number")
    if sku_i < 0:
        wb.close()
        raise ValueError(f"ASD Excel missing 'sku' column. Headers: {header}")

    out: List[Dict[str, str]] = []
    for row in rows_iter:
        if not row:
            continue
        sku = str(row[sku_i] or "").strip() if sku_i < len(row) else ""
        if not sku:
            continue
        out.append(
            {
                "brand_name": str(row[brand_i] or "").strip() if brand_i >= 0 and brand_i < len(row) else "",
                "sku": normalize_sku(sku),
                "brand_url": str(row[url_i] or "").strip() if url_i >= 0 and url_i < len(row) else "",
                "listing_page_url": str(row[page_url_i] or "").strip() if page_url_i >= 0 and page_url_i < len(row) else "",
                "page_number": str(row[page_num_i] or "").strip() if page_num_i >= 0 and page_num_i < len(row) else "",
            }
        )
    wb.close()
    return out


def load_asm_rows(client: gspread.Client) -> Tuple[List[str], List[List[str]]]:
    ss = safe_gs_call(client.open_by_key, ASM_SPREADSHEET_ID)
    ws = safe_gs_call(ss.worksheet, ASM_PRODUCTS_TAB)
    values = safe_gs_call(ws.get_all_values)
    if not values:
        return [], []
    return values[0], values[1:]


def header_index(header: List[str], name: str) -> int:
    target = name.strip().lower()
    for i, h in enumerate(header):
        if str(h).strip().lower() == target:
            return i
    return -1


def cell(row: List[str], idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def build_asm_records(header: List[str], data_rows: List[List[str]]) -> List[Dict[str, str]]:
    pid_i = header_index(header, "product_id")
    ref_i = header_index(header, "Reference")
    brand_i = header_index(header, "Brand1")
    title_i = header_index(header, "Title")
    price_i = header_index(header, "Price")
    avail_i = header_index(header, "Availability")
    url_i = header_index(header, "url")

    records: List[Dict[str, str]] = []
    for row in data_rows:
        ref = cell(row, ref_i)
        if not ref:
            continue
        records.append(
            {
                "product_id": cell(row, pid_i),
                "Reference": ref,
                "sku": normalize_sku(ref),
                "Brand1": cell(row, brand_i),
                "Title": cell(row, title_i),
                "Price": cell(row, price_i),
                "Availability": cell(row, avail_i),
                "url": cell(row, url_i),
            }
        )
    return records


def unique_missing(
    records: List[Dict[str, str]],
    other_keys: Set[str],
) -> List[Dict[str, str]]:
    """Keep one row per lowercase SKU that exists in records but not in other_keys."""
    seen: Set[str] = set()
    out: List[Dict[str, str]] = []
    for rec in records:
        key = rec.get("sku") or ""
        if not key or key in other_keys or key in seen:
            continue
        seen.add(key)
        out.append(rec)
    return out


def compare_sets(
    asm_records: List[Dict[str, str]],
    asd_records: List[Dict[str, str]],
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], Set[str], Set[str]]:
    asm_keys = {r["sku"] for r in asm_records if r.get("sku")}
    asd_keys = {r["sku"] for r in asd_records if r.get("sku")}

    asm_not = unique_missing(asm_records, asd_keys)
    asd_not = unique_missing(asd_records, asm_keys)
    return asm_not, asd_not, asm_keys, asd_keys


def rows_from_dicts(headers: List[str], records: List[Dict[str, str]]) -> List[List[str]]:
    return [[r.get(h, "") for h in headers] for r in records]


def upsert_tab(
    spreadsheet: gspread.Spreadsheet,
    title: str,
    headers: List[str],
    rows: List[List[str]],
) -> gspread.Worksheet:
    try:
        ws = safe_gs_call(spreadsheet.worksheet, title)
    except gspread.WorksheetNotFound:
        ws = safe_gs_call(
            spreadsheet.add_worksheet,
            title=title,
            rows=max(1000, len(rows) + 10),
            cols=max(len(headers) + 2, 10),
        )
    safe_gs_call(ws.clear)
    safe_gs_call(ws.append_row, headers, value_input_option="USER_ENTERED")
    if rows:
        batch = 500
        for i in range(0, len(rows), batch):
            safe_gs_call(ws.append_rows, rows[i : i + batch], value_input_option="USER_ENTERED")
    return ws


def write_csv(path: str, headers: List[str], rows: List[List[str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare ASM Google Sheet SKUs vs ASD distribution Excel.")
    parser.add_argument("--asd-xlsx", default="", help="Path to ASD Excel (default: latest distribution_brand_skus*.xlsx)")
    parser.add_argument("--write-csv", action="store_true", help="Also write CSV files locally")
    args = parser.parse_args()

    asd_path = find_latest_asd_xlsx(args.asd_xlsx)
    print(f"ASD Excel: {asd_path}")

    asd_records = load_asd_rows(asd_path)
    print(f"ASD rows: {len(asd_records)} ({len({r['sku'] for r in asd_records})} unique lowercase SKUs)")

    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)

    asm_header, asm_data = load_asm_rows(client)
    if not asm_header:
        print("[ERROR] ASM products sheet is empty.")
        return 1
    asm_records = build_asm_records(asm_header, asm_data)
    print(f"ASM rows: {len(asm_records)} ({len({r['sku'] for r in asm_records})} unique lowercase SKUs)")

    asm_not, asd_not, asm_keys, asd_keys = compare_sets(asm_records, asd_records)
    overlap = asm_keys & asd_keys
    print(f"Matched SKUs (lowercase): {len(overlap)}")
    print(f"ASM only (unique SKUs): {len(asm_not)}")
    print(f"ASD only (unique SKUs): {len(asd_not)}")

    asm_rows = rows_from_dicts(ASM_NOT_HEADERS, asm_not)
    asd_rows = rows_from_dicts(ASD_NOT_HEADERS, asd_not)

    ss = safe_gs_call(client.open_by_key, ASM_SPREADSHEET_ID)
    upsert_tab(ss, TAB_ASM_NOT_IN_ASD, ASM_NOT_HEADERS, asm_rows)
    upsert_tab(ss, TAB_ASD_NOT_IN_ASM, ASD_NOT_HEADERS, asd_rows)

    sheet_url = f"https://docs.google.com/spreadsheets/d/{ASM_SPREADSHEET_ID}/edit"
    print(f"\n[OK] Wrote tab '{TAB_ASM_NOT_IN_ASD}' ({len(asm_rows)} rows)")
    print(f"[OK] Wrote tab '{TAB_ASD_NOT_IN_ASM}' ({len(asd_rows)} rows)")
    print(f"Sheet: {sheet_url}")

    if args.write_csv:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        asm_csv = os.path.join(SCRIPT_DIR, f"asm_not_in_asd_{ts}.csv")
        asd_csv = os.path.join(SCRIPT_DIR, f"asd_not_in_asm_{ts}.csv")
        write_csv(asm_csv, ASM_NOT_HEADERS, asm_rows)
        write_csv(asd_csv, ASD_NOT_HEADERS, asd_rows)
        print(f"[OK] CSV: {asm_csv}")
        print(f"[OK] CSV: {asd_csv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
